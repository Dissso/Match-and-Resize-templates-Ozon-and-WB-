import shutil
import tkinter as tk
import tkinter.ttk as ttk
from PIL import Image, ImageTk
import os
import json
import glob
import openpyxl
from tkinter import filedialog
import requests
from io import BytesIO
import re
import datetime
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity

# Create a Tkinter window
root = tk.Tk()
root.title("Matching templates")
root.geometry("800x600")  # Set the window size to 800x600 pixels

# Add an image to the window
image_path = "ozon-1-3.jpg"
image = Image.open(image_path)
photo = ImageTk.PhotoImage(image)
label = tk.Label(root, image=photo)
label.pack()

# Create a frame to hold the buttons
frame = tk.Frame(root, bg="white")  # Create a frame with a white background
frame.place(relx=0.5, rely=0.5, anchor=tk.CENTER)  # Place the frame in the center of the window


templates_folder = 'Templates_WB'
templates = [f for f in os.listdir(templates_folder) if f.endswith('.xlsx')]

print(templates)  # should print the list of template files

# Create a label to prompt the user to select a template
template_label = tk.Label(root, text="Выберите категорию WB:")
template_label.place(relx=0.5, rely=0.25, anchor=tk.CENTER)

# Load the JSON data from the file
with open('WB_Categories.json', 'r', encoding='utf-8') as f:
    categories = json.load(f)

# Create a frame to hold the search input and output text
main_frame = tk.Frame(root)
main_frame.place(relx=0.5, rely=0.5, anchor=tk.CENTER)

# Create a frame for the search input
search_frame = tk.Frame(main_frame)
search_frame.pack(side="top")

# # Create a label and entry for the search query
search_query = tk.StringVar()

search_entry = tk.Entry(search_frame, width=40, textvariable=search_query)
search_entry.insert(0, "Введите запрос для поиска категории:")
search_entry.bind("<FocusIn>", lambda event: search_entry.delete(0, tk.END))
search_entry.pack(side="top")

# Create a listbox to display the search results
search_results = tk.Listbox(search_frame, width=40)
search_results.pack(side="top")
search_results.bind("<<ListboxSelect>>", lambda event: update_selected_category(event, search_results, selected_category_label))

# Create a label to display the selected category
selected_label = tk.Label(search_frame, text="Выбрана категория:")
selected_label.pack(side="top")
selected_category_label = tk.Label(search_frame, text="")
selected_category_label.pack(side="top")

# Create a Text widget to display the output
output_text = tk.Text(main_frame, width=40, height=10)
output_text.pack(side="top")

selected_category = ""
# Function to update the selected category label
def update_selected_category(event, listbox, label):
    global selected_category
    selection = listbox.curselection()
    if selection:
        selected_index = selection[0]
        selected_category = listbox.get(selected_index)
        for category in categories["data"]:
            if selected_category in [category["title"]] + [child["title"] for child in category.get("children", [])]:
                label.config(text=f"{category['title']} > {selected_category}")
                print(category['title'])
                print(selected_category)
    else:
        selected_category_label.config(text=label.cget("text"))  # Update the selected_category_label text
# Function to search categories
def search_categories(*args):
    query = search_query.get()
    search_results.delete(0, tk.END)  # Clear the Listbox
    if query:
        # Extract the category titles
        category_titles = []
        for category in categories["data"]:
            category_titles.append(category["title"])
            for child in category.get("children", []):
                category_titles.append(child["title"])

        # Filter categories that match the query
        matching_categories = [category for category in category_titles if query.lower() in category.lower()]
        for category in matching_categories:
            search_results.insert(tk.END, category)

# Bind the search query to the search function
search_query.trace("w", search_categories)



def load_template():
    global selected_category
    # wb = None
    category_value = selected_category_label.cget("text")
    parts = category_value.split("> ")
    if len(parts) > 1:
        category_product = parts[1]
    else:
        category_product = None
    # Извлечение основной категории и подкатегории
    main_category = parts[0].strip()  # Основная категория
    sub_category = parts[1].strip() if len(parts) > 1 else None  # Подкатегория

    print(f"Main Category: '{main_category}'")  # Отладочный вывод
    print(f"Sub Category: '{sub_category}'")  # Отладочный вывод
    for category in categories["data"]:
        if category_value.startswith(f"{category['title']} > "):
            selected_category = category_product
            break
    else:
        selected_category = None

    # Проверка, нашли ли мы категорию
    if selected_category:
        for template in templates:
            # Полное соответствие с учетом расширения файла
            if template == f"{selected_category}.xlsx":  # Предполагаем, что шаблоны имеют расширение .xlsx
                template_path = os.path.join(templates_folder, template)
                wb = openpyxl.load_workbook(template_path)
                print(f"Загружен темплейт: {template} категория {category_value}")
                selected_template = template  # Определяем и присваиваем selected_template
                break
        else:
            output_text.delete(1.0, tk.END)  # clear the Text widget
            output_text.insert(tk.END, "No template loaded")
            print("No matching template found")
            selected_template = None
            wb = openpyxl.Workbook()
    else:
        print("No category selected")
        selected_template = None
        wb = openpyxl.Workbook()

    if wb is None:
        wb = openpyxl.Workbook()

    sheet_dst = wb.active
    row = next(sheet_dst.iter_rows(min_row=3, max_row=3))
    characteristics = [str(cell.value) for cell in row[1:] if cell.value is not None]
    print('Characteristics:',characteristics)
    num_characteristics = len(characteristics)
    print(num_characteristics)

    # Insert the loaded template message into the output Text widget
    output_text.delete(1.0, tk.END)  # clear the Text widget
    if selected_template:
        output_text.config(state='normal')  # Re-enable the Text widget
        output_text.delete(1.0, tk.END)  # Clear the Text widget
        output_text.insert(tk.END, f"Загружен темплейт: {selected_template}")
    else:
        output_text.insert(tk.END, "Ошибка! No template loaded \n")
        output_text.insert(tk.END, "Попробуйте выбранную категорию WB пересохранить а папке \ntemplates_WB")
    print('Wb1:', wb)

    return wb, characteristics
# Create a button to load the template
load_button = tk.Button(search_frame, text="Загрузите выбранный темплейт WB", command=load_template)
load_button.pack(side="top")

def select_file1():
    file_path = filedialog.askopenfilename(title="Выберите Excel file (with data)",
                                           filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        filename1 = os.path.basename(file_path)  # Extract the filename from the full path
        filename = os.path.splitext(filename1)[0]
        wb1 = openpyxl.load_workbook(file_path)
        sheet1 = wb1.worksheets[4]
        # select the 5th sheet (index 4)
        comparison_characteristics = []

        for row in sheet1.iter_rows(min_row=2, max_row=2):
            for idx, cell in enumerate(row):
                if idx > 0:  # Skip the first cell
                    if cell.value is not None:
                        comparison_characteristics.append(str(cell.value))
        print(len(comparison_characteristics))

        # Insert a new first row with empty values
        sheet1.insert_rows(1)  # Insert a new row at index 1 (i.e., the first row)

        # Set the values of the first row to empty strings
        for cell in sheet1[1]:
            cell.value = ""


        for cell in sheet1[2]:
            print(cell.value)

        # Define the photo columns
        photo_cols = ['Ссылка на главное фото*', 'Ссылки на дополнительные фото', 'Ссылки на фото 360']
        photo_cols_idx = []
        for col in photo_cols:
            try:
                col_idx = next(i for i, cell in enumerate(sheet1[3], start=1) if cell.value == col)
                photo_cols_idx.append(col_idx)
            except StopIteration:
                raise ValueError(f"Column '{col}' not found in the sheet")

            # Iterate over the rows, skipping empty rows
            for row_idx, row in enumerate(sheet1.iter_rows(values_only=True), start=4):
                if any(cell for cell in row):  # Check if the row is not empty
                    merged_photo_value = []
                    for col_idx in photo_cols_idx:
                        cell_value = sheet1.cell(row=row_idx, column=col_idx).value
                        if cell_value:
                            cell_value = cell_value.replace('\n', ';')  # Remove newline characters
                            merged_photo_value.extend(str(cell_value).split(';'))
                            # print(f'Merged_photo {merged_photo_value}')
                    sheet1.cell(row=row_idx, column=photo_cols_idx[0]).value = ';'.join(merged_photo_value)
                    # Clear the other photo columns
                    for col_idx in photo_cols_idx[1:]:
                        sheet1.cell(row=row_idx, column=col_idx).value = None


        return wb1, sheet1, filename, comparison_characteristics

def compare_characteristics(characteristics,comparison_characteristics):
    global selected_category_label
    global selected_category
    # Get the characteristics from both files
    template_characteristics = characteristics
    comparison_characteristics = comparison_characteristics
    print('Template_characteristics:', len(template_characteristics))
    print('Comparison_characteristics:', len(comparison_characteristics))

    vectorizer = TfidfVectorizer()

    # Initialize an empty dictionary to store the maximum cosine similarity values
    max_similarity_values = {}

    printed_keys = set()
    # Iterate over the keys in dict1
    for key1 in template_characteristics:
        # Initialize the maximum similarity value for the current key
        max_similarity = 0
        max_similarity_key = ""

        # Iterate over the keys in dict2
        for key2 in comparison_characteristics:
            # Create a list of single-key strings
            keys = [key1, key2]

            # Fit and transform the keys
            tfidf_matrix = vectorizer.fit_transform(keys)
            tfidf_array = tfidf_matrix.toarray()

            # Calculate cosine similarity
            similarity = cosine_similarity(tfidf_array)[0][1]

            # Update maximum similarity if found a new maximum
            if similarity > max_similarity:
                max_similarity = similarity
                max_similarity_key = key2
            # If the similarity is the same, choose the one that comes first alphabetically
            elif similarity == max_similarity and key2 < max_similarity_key:
                max_similarity_key = key2

        # Store the maximum similarity value and key in the dictionary
        max_similarity_values[key1] = (max_similarity_key, max_similarity)

    output_dict = {}
    count = 1
    for key1, (key2, similarity) in max_similarity_values.items():

        if similarity > 0.3 and key2 not in printed_keys:
            print(f"{count} Косинусное сходство между хар-ми '{key1}' и '{key2}': {similarity:.2f}")
            printed_keys.add(key2)
            count += 1
            output_dict[key1] = key2

    # Вывод результата
    print(output_dict)

    # # Check if the JSON file exists in the Column_maps directory
    json_file_path = f'Column_maps/column_map_{selected_category}.json'

    predefined_column_map = {
        "Баркоды": "Артикул*",
        "Артикул производителя": "Партномер*",
        "Наименование": "Название товара",
        "Цена": "Цена, руб.*",
        "Категория продавца": "Тип*",
        "Ставка НДС": "НДС, %*",
        "Вес с упаковкой (кг)": "Вес в упаковке, г*",
        "Ширина упаковки": "Ширина упаковки, мм*",
        "Высота упаковки": "Высота упаковки, мм*",
        "Длина упаковки": "Длина упаковки, мм*",
        "Фото": "Ссылка на главное фото*",
        "Бренд": "Бренд*",
        "Комплектация": "Комплектация",
        "ОЕМ номер": "OEM-номер",
        "Страна производства": "Страна-изготовитель",
        "Вес товара без упаковки (г)": "Вес товара, г*",
        "Размер": "Размер без подставки (ШxВxГ), мм*",
        "Гарантийный срок": "Гарантийный срок*",
        "Описание": "Аннотация",
        "Артикул продавца": "Ссылки на фото 360"
    }

    for key, value in output_dict.items():

        if key not in predefined_column_map:

            if value not in predefined_column_map.values():
                predefined_column_map[key] = value
        else:

            if predefined_column_map[key] != value:
                pass

    print('Обновленный predefined_column_map:', predefined_column_map)
    print(f'Количество совпадений характеристик: {len(predefined_column_map)}')

    column_map = {}

    for key in template_characteristics:

        if key in predefined_column_map:

            column_map[key] = predefined_column_map[key]
        else:

            column_map[key] = ''



    with open(f'Column_maps/column_map_{selected_category}.json', 'w', encoding='utf-8') as f:

        json.dump(column_map, f, ensure_ascii=False, indent=4)

    for i, (key, value) in enumerate(column_map.items(), start=1):
        print(f"{i} {key} - {value}")

    # Open the Column Map Editor window
    open_column_map_editor(column_map, comparison_characteristics)

def open_column_map_editor(predefined_column_map, comparison_characteristics):
    global combo_fields
    global selected_category
    column_map_editor = tk.Toplevel(root)
    column_map_editor.title("Column Map Editor")



    combo_fields = {}

    update_button_frame = tk.Frame(column_map_editor)
    update_button_frame.grid(row=0, column=0, columnspan=2, sticky="ew")

    column_map = {}  # Initialize the column map

    update_button = tk.Button(update_button_frame, text="Update Column Map",
                              command=lambda: column_map.update(update_column_map(combo_fields, column_map)))
    update_button.pack(fill="x")

    # Create a frame to hold the list of characteristics
    char_frame = tk.Frame(column_map_editor)
    char_frame.grid(row=1, column=0, columnspan=2, sticky="nsew")


    # Create a canvas to hold the list of characteristics
    canvas = tk.Canvas(char_frame)
    canvas.pack(side="left", fill="both", expand=True)

    scrollbar = tk.Scrollbar(char_frame, orient="vertical", command=canvas.yview)
    scrollbar.pack(side="right", fill="y")
    canvas.config(yscrollcommand=scrollbar.set)

    # Add binding to canvas to respond to mouse wheel events
    canvas.bind_all("<MouseWheel>", lambda event: canvas.yview_scroll(-1 if event.delta > 0 else 1, "units"))


    # Create a frame to hold the list of characteristics inside the canvas
    char_inner_frame = tk.Frame(canvas)
    canvas.create_window((0, 0), window=char_inner_frame, anchor="nw")



    row_num = 1
    for key, value in predefined_column_map.items():
        label = tk.Label(char_inner_frame, text=key)
        label.grid(row=row_num, column=0)
        combo = ttk.Combobox(char_inner_frame, width=40, values=comparison_characteristics)
        combo.set(value)
        combo.grid(row=row_num, column=1)
        combo_fields[key] = combo
        # print('Combo_fields:',combo_fields)
        row_num += 1


    # Update the canvas to show the entire frame
    char_inner_frame.update_idletasks()
    canvas.configure(scrollregion=canvas.bbox("all"))

    req_width = char_inner_frame.winfo_reqwidth()
    req_height = char_inner_frame.winfo_reqheight()
    column_map_editor.geometry(f"{req_width + 20}x{req_height}")

    # Set the window to resize dynamically
    column_map_editor.rowconfigure(1, weight=1)
    column_map_editor.columnconfigure(0, weight=1)

    column_map_editor.protocol("WM_DELETE_WINDOW", lambda: (
    combo_fields.clear() if len(combo_fields) == 0 else None, column_map_editor.destroy()))



def update_column_map(combo_fields, column_map):
    global selected_category  # access the global variable
    # column_map = {}
    for key, combo in combo_fields.items():
        new_value = combo.get()
        column_map[key] = new_value
    print("Updated column map:", column_map)

    with open(f'Column_maps/column_map_{selected_category}.json', 'w', encoding='utf-8') as f:
        f.write("column_map = ")
        json.dump(column_map, f, ensure_ascii=False, indent=4)
    return column_map


def on_select_file1_click():
    global wb1, sheet1, filename, comparison_characteristics, characteristics
    wb, characteristics = load_template()
    wb1, sheet1, filename, comparison_characteristics = select_file1()
    print("select_file1() returned:", wb1, sheet1, filename, comparison_characteristics)
    # You can add some code here to display the selected file information
    output_text.config(state='normal')  # Re-enable the Text widget
    output_text.delete(1.0, tk.END)  # Clear the Text widget
    output_text.insert(tk.END, f"Выбран шаблон Озон: {filename}\n")  # Insert the selected file name
    compare_characteristics(characteristics, comparison_characteristics)
    print('Comparison_characteristics:', comparison_characteristics)

def match_files(wb, sheet1, wb_dst, sheet_dst, filename, category_product, column_map):

    column_map = {v: k for k, v in column_map.items()}

    print('Column_map:', column_map)

    wb = load_template()
    print('Sheet_dst:', sheet_dst)
    print('Wb:', wb)



    match_count = 0
    articul_values_index = ''
    articul_values = []
    num_cols = sum(1 for col in sheet1.iter_cols(values_only=True) if any(cell is not None for cell in col))
    print(f'Всего хар-ик со значениями: {num_cols}')
    # Then, write the data from sheet1 to row 5 and onwards
    total_cols = sheet1.max_column  # Define total_cols
    for col_idx, col in enumerate(sheet1.iter_cols(values_only=True), start=1):
        output_text.insert(tk.END, f"Processing columns: {col_idx}/{total_cols} ({col_idx / total_cols * 100:.2f}%)\n")
        output_text.see(tk.END)  # Scroll to the end of the text widget
        output_text.update_idletasks()  # Update the text widget
        col_name = sheet1.cell(row=3, column=col_idx).value
        col_name = sheet1.cell(row=3, column=col_idx).value
        if col_name == 'Артикул*':
            articul_values = list(col)[4:]
            print(articul_values)
            articul_values_index = 0
        if col_name in column_map:
            col2_name = column_map[col_name]
            print(f"Mapping {col_name} to {col2_name}")
            # Find the corresponding column in the second file
            for col2_idx, col2 in enumerate(sheet_dst.iter_cols(values_only=True), start=1):
                if sheet_dst.cell(row=3, column=col2_idx).value == col2_name:
                    print(f"Found column {col2_name} in second file")
                    name_col_idx = next(i for i, cell in enumerate(sheet1[3], start=1) if cell.value == 'Название товара')
                    # Write the data from sheet1 to sheet_dst
                    for row_idx in range(5, sheet1.max_row + 1):
                        print(row_idx)
                        # Write the data from sheet1 to sheet_dst
                        name_value = sheet1.cell(row=row_idx, column=name_col_idx).value
                        if name_value:  # Check if the value in the 'Наименование' column is not empty
                            cell_value = sheet1.cell(row=row_idx, column=col_idx).value

                            if col2_name == 'Артикул продавца':
                                if articul_values_index < len(articul_values) and articul_values[articul_values_index] is not None:
                                    value = articul_values[articul_values_index]
                                    sheet_dst.cell(row=row_idx, column=col2_idx).value = value
                                    print(f"Writing value {value} to row {row_idx}, column {col2_idx}")
                                    print(f"Articul_values: {value}")
                                    articul_values_index += 1  # increment the index for the next row

                            else:
                                if col2_name == 'Баркоды':
                                    cell_value = str(cell_value) + '-MK'
                                if col2_name == 'Вес с упаковкой (кг)':
                                    if cell_value is not None:
                                        cell_value = float(cell_value) / 1000  # Convert grams to kilograms
                                if col2_name in ['Ширина упаковки', 'Высота упаковки', 'Длина упаковки']:
                                    if cell_value is not None:
                                        cell_value = cell_value / 10  # Convert mm to cm
                                print(f"Writing value {cell_value} to row {row_idx}, column {col2_idx}")
                                sheet_dst.cell(row=row_idx, column=col2_idx).value = cell_value
                            if col2_name == 'Категория продавца':

                                sheet_dst.cell(row=row_idx, column=col2_idx).value = category_product
                    match_count += 1
                    break

    # Save the updated second file to a new Excel file
    print("Saving workbook to Match_Ozon&WB.xlsx")
    wb_dst.save(f'{category_product}_Match.xlsx')
    print("Workbook saved successfully")
    result = f"Совпадений хар-ик: {match_count} \nРезультат: {category_product}_Match.xlsx"  # Replace with the actual result
    print('Wb_dst:',wb_dst)
    return sheet_dst

def on_match_files_click():
    global combo_fields
    global selected_category
    wb_dst, characteristics = load_template()
    print('Wb_dst:', wb_dst)
    sheet_dst = wb_dst.active
    print(sheet_dst)
    print("wb1:", wb1)
    print("sheet1:", sheet1)
    column_map = {}  # Initialize column_map as an empty dictionary
    column_map = update_column_map(combo_fields, column_map)
    result = match_files(wb1, sheet1, wb_dst, sheet_dst, filename, selected_category_label.cget("text").split(">")[1].strip(), column_map)

    output_text.config(state='normal')  # Re-enable the Text widget
    output_text.delete(1.0, tk.END)  # Clear the Text widget
    output_text.insert(tk.END, result)  # Insert the result into the Text widget
    # Create a button to execute the on_resize_image_click function
    resize_button = tk.Button(search_frame, text="Изменить размер фото", command=lambda: resize_image(wb_dst))
    resize_button.pack(side="top")


# Create a button to execute the on_select_file1_click function
select_button = tk.Button(search_frame, text="Выберите шаблон Ozon", command=on_select_file1_click)
select_button.pack(side="top")

# Create a button to execute the on_match_files_click function
match_button = tk.Button(search_frame, text="Сопоставление карточек", command=on_match_files_click)
match_button.pack(side="top")



def resize_image(wb_dst):
    main_folder = f'Img_{selected_category}'

    if os.path.exists(main_folder):
        shutil.rmtree(main_folder)
        print(f"Содержимое папки '{main_folder}' удалено успешно!")
    sheet_dst2 = wb_dst.active
    # Iterate over the rows in the "Фото" column
    data = [list(row) for row in sheet_dst2.iter_rows(values_only=True)]
    total_rows = len(data)  # Define total_rows
    image_urls = ''
    for row_idx, row in enumerate(sheet_dst2.iter_rows(values_only=True), start=1):
        if row_idx >= 5:  # start from the 3rd row
            output_text.insert(tk.END,
                               f"{row_idx - 4} Карточка обработана: {row_idx - 4}/{total_rows} ({row_idx / total_rows * 100:.2f}%)\n")
            output_text.see(tk.END)  # Scroll to the end of the text widget
            output_text.update_idletasks()  # Update the text widget
            row = list(sheet_dst2.iter_rows(values_only=True))[row_idx - 1]
            image_url = row[7]
            print(image_url)
            if row[7] is not None:
               image_urls = row[7].split(';')
            names = row[3]
            names = re.sub(r'[<>:"/\\|?*,. ]', '_', names)
            print(names)
            print(image_urls)
            new_image_url = resize_image_url(image_urls, names)

    output_text.config(state='normal')  # Re-enable the Text widget
    output_text.insert(tk.END, "Обработка фотографий завершена!\n")
    output_text.see(tk.END)  # Scroll to the end of the text widget
    output_text.update_idletasks()  # Update the text widget



def resize_image_url(image_urls, names):
    print('Image_urls:', image_urls)

    main_folder = f'Img_{selected_category}'

    try:
        os.mkdir(main_folder)
        print(f"Папка '{main_folder}' создана успешно!")
    except FileExistsError:
        print(f"Папка '{main_folder}' уже существует!")
    except OSError as e:
        print(f"Ошибка создания папки: {e}")


    try:
        os.makedirs(os.path.join('main_folder', names))
    except FileExistsError:
        pass


    for url in image_urls:
        image_url = url.strip()  # удалить пробелы в начале и конце строки
        if image_url and image_url.startswith(("http://", "https://")):
            print(image_url)
            filename = image_url.split('/')[-1]

            filename1 = filename.rsplit('.', 1)[0]

            image_data = requests.get(image_url).content

            image = Image.open(BytesIO(image_data))

            # Resize the image if necessary
            width, height = image.size
            print(f'{image_url} Изображение Разрешение: {image.size}')
            if width < 900 or height < 1200:
                # Calculate the new height to maintain 3:4 aspect ratio
                new_height = int(image.width * 4 / 3)

                # Calculate the padding needed to add a white background
                padding = (new_height - image.height) // 2

                # Create a new image with white background
                new_image = Image.new('RGB', (image.width, new_height), (255, 255, 255))

                # Paste the original image onto the new image
                new_image.paste(image, (0, padding))

                # Replace the original image with the new image
                image = new_image

                # Calculate the new width and height
                aspect_ratio = image.width / image.height
                if image.width < 900:
                    new_width = 900
                    new_height = int(new_width / aspect_ratio)
                    if new_height < 1200:
                        new_height = 1200
                        new_width = int(new_height * aspect_ratio)
                elif image.height < 1200 or image.width < 900:
                    if image.height < 1200:
                        new_height = 1200
                        new_width = int(new_height * aspect_ratio)
                    else:
                        new_width = 900
                        new_height = int(new_width / aspect_ratio)
                else:
                    new_width = image.width
                    new_height = image.height
                # Resize the image
                image = image.resize((new_width, new_height))
                print(f'Новое разрешение файла Доп.фото:{image.size}')
                image_filename = f"{filename1}_new.jpg"
            else:
                image = image.resize((width, height))
                image_filename = f"{filename1}.jpg"

            # Convert the image to a bytes object
            image_bytes = BytesIO()
            image.save(image_bytes, format='JPEG')
            image_bytes.seek(0)

            os.makedirs(os.path.join('Img_' + selected_category, names), exist_ok=True)
            with open(os.path.join('Img_' + selected_category, names, image_filename), 'wb') as f:
                f.write(image_bytes.read())
        else:
            print("Неправильный URL")





root.mainloop()
